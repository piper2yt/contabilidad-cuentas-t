import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import datetime, date
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── CONFIG ───────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="ContaT · Cuentas T Contables",
    page_icon="📒",
    layout="wide",
    initial_sidebar_state="expanded",
)

EXCEL_FILE = "movimientos_contables.xlsx"
CUENTAS_DEFAULT = ["Activos", "Pasivos", "Capital", "Ingresos", "Gastos"]
MOVIMIENTOS = ["CARGOS", "ABONOS"]

NATURALEZA = {
    "Activos": "deudora",
    "Gastos": "deudora",
    "Pasivos": "acreedora",
    "Capital": "acreedora",
    "Ingresos": "acreedora",
}

# ─── STYLES ───────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Serif+Display&family=DM+Sans:wght@300;400;500;600&display=swap');

html, body, [class*="css"] {
    font-family: 'DM Sans', sans-serif;
}

h1, h2, h3 { font-family: 'DM Serif Display', serif; }

.stApp { background: #0f1117; color: #e8e3d5; }

section[data-testid="stSidebar"] {
    background: #161922;
    border-right: 1px solid #2a2d3a;
}

.metric-card {
    background: linear-gradient(135deg, #1a1e2e 0%, #1f2438 100%);
    border: 1px solid #2e3347;
    border-radius: 12px;
    padding: 1.2rem 1.4rem;
    margin-bottom: 0.8rem;
}
.metric-card .label {
    font-size: 0.72rem;
    letter-spacing: 0.12em;
    text-transform: uppercase;
    color: #6b7280;
    margin-bottom: 0.3rem;
}
.metric-card .value {
    font-size: 1.8rem;
    font-family: 'DM Serif Display', serif;
    font-weight: 400;
}
.positive { color: #34d399; }
.negative { color: #f87171; }
.neutral  { color: #93c5fd; }

.cuenta-t {
    background: #161922;
    border: 1px solid #2e3347;
    border-radius: 10px;
    overflow: hidden;
    margin-bottom: 1rem;
}
.cuenta-t-header {
    background: linear-gradient(90deg, #1e3a5f, #1a2d4a);
    padding: 0.6rem 1rem;
    font-family: 'DM Serif Display', serif;
    font-size: 1rem;
    color: #93c5fd;
    border-bottom: 1px solid #2e3347;
    display: flex;
    justify-content: space-between;
    align-items: center;
}
.cuenta-t-body {
    display: grid;
    grid-template-columns: 1fr 1px 1fr;
    min-height: 100px;
}
.cuenta-t-col {
    padding: 0.8rem;
}
.cuenta-t-divider {
    background: #2e3347;
}
.cargo-label { color: #f87171; font-size:0.7rem; letter-spacing:0.1em; text-transform:uppercase; }
.abono-label { color: #34d399; font-size:0.7rem; letter-spacing:0.1em; text-transform:uppercase; }
.t-amount { font-size: 0.9rem; color: #e8e3d5; padding: 2px 0; }

.pill {
    display: inline-block;
    padding: 2px 10px;
    border-radius: 20px;
    font-size: 0.72rem;
    font-weight: 600;
    letter-spacing: 0.05em;
}
.pill-cargo { background: #3b1c1c; color: #f87171; }
.pill-abono { background: #1a3b2e; color: #34d399; }

.esf-ok {
    background: linear-gradient(90deg, #1a3b2e, #1f2438);
    border: 1px solid #34d399;
    border-radius: 8px;
    padding: 0.9rem 1.2rem;
    color: #34d399;
    font-weight: 600;
}
.esf-fail {
    background: linear-gradient(90deg, #3b1c1c, #1f2438);
    border: 1px solid #f87171;
    border-radius: 8px;
    padding: 0.9rem 1.2rem;
    color: #f87171;
    font-weight: 600;
}

div[data-testid="stForm"] {
    background: #161922;
    border: 1px solid #2e3347;
    border-radius: 12px;
    padding: 1.4rem;
}

.stButton > button {
    background: linear-gradient(135deg, #2563eb, #1d4ed8) !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    font-family: 'DM Sans', sans-serif !important;
    transition: all 0.2s !important;
}
.stButton > button:hover {
    background: linear-gradient(135deg, #3b82f6, #2563eb) !important;
    transform: translateY(-1px);
    box-shadow: 0 4px 12px rgba(37,99,235,0.4) !important;
}

.stSelectbox > div > div,
.stNumberInput > div > div > input,
.stTextInput > div > div > input,
.stDateInput > div > div > input {
    background: #1f2438 !important;
    border: 1px solid #2e3347 !important;
    border-radius: 8px !important;
    color: #e8e3d5 !important;
}

.tab-header {
    font-size: 0.75rem;
    letter-spacing: 0.15em;
    text-transform: uppercase;
    color: #6b7280;
    margin-bottom: 1.5rem;
    padding-bottom: 0.5rem;
    border-bottom: 1px solid #2e3347;
}

.stDataFrame { border-radius: 10px; overflow: hidden; }
</style>
""", unsafe_allow_html=True)


# ─── EXCEL HELPERS ────────────────────────────────────────────────────────────
SEED_FILE = "movimientos_contables.xlsx"   # archivo de prueba en el repo
WORK_FILE = "movimientos_trabajo.xlsx"      # archivo de trabajo en tiempo de ejecución

DATOS_SEMILLA = [
    ("2025-01-03","Activos","CARGOS",50000,1,2025,"Efectivo aportado por el dueño"),
    ("2025-01-03","Capital","ABONOS",50000,1,2025,"Capital inicial del dueño"),
    ("2025-01-05","Activos","CARGOS",20000,1,2025,"Préstamo bancario recibido"),
    ("2025-01-05","Pasivos","ABONOS",20000,1,2025,"Préstamo bancario por pagar"),
    ("2025-01-07","Gastos", "CARGOS",15000,1,2025,"Compra de inventario inicial"),
    ("2025-01-07","Activos","ABONOS",15000,1,2025,"Pago inventario en efectivo"),
    ("2025-01-08","Activos","CARGOS",12000,1,2025,"Equipo de cómputo adquirido"),
    ("2025-01-08","Activos","ABONOS",12000,1,2025,"Pago equipo en efectivo"),
    ("2025-01-10","Activos","CARGOS", 8500,1,2025,"Cobro ventas semana 1"),
    ("2025-01-10","Ingresos","ABONOS",8500,1,2025,"Ventas semana 1 enero"),
    ("2025-01-17","Activos","CARGOS", 9200,1,2025,"Cobro ventas semana 2"),
    ("2025-01-17","Ingresos","ABONOS",9200,1,2025,"Ventas semana 2 enero"),
    ("2025-01-24","Activos","CARGOS", 7800,1,2025,"Cobro ventas semana 3"),
    ("2025-01-24","Ingresos","ABONOS",7800,1,2025,"Ventas semana 3 enero"),
    ("2025-01-31","Activos","CARGOS", 6500,1,2025,"Cobro ventas semana 4"),
    ("2025-01-31","Ingresos","ABONOS",6500,1,2025,"Ventas semana 4 enero"),
    ("2025-01-31","Gastos", "CARGOS", 5000,1,2025,"Sueldo empleado enero"),
    ("2025-01-31","Activos","ABONOS", 5000,1,2025,"Pago sueldo enero"),
    ("2025-01-31","Gastos", "CARGOS", 3500,1,2025,"Renta local enero"),
    ("2025-01-31","Activos","ABONOS", 3500,1,2025,"Pago renta enero"),
    ("2025-01-31","Gastos", "CARGOS",  800,1,2025,"Servicios luz y agua enero"),
    ("2025-01-31","Activos","ABONOS",  800,1,2025,"Pago servicios enero"),
    ("2025-02-03","Pasivos","CARGOS", 8000,2,2025,"Abono parcial a préstamo bancario"),
    ("2025-02-03","Activos","ABONOS", 8000,2,2025,"Pago préstamo con efectivo"),
    ("2025-02-05","Gastos", "CARGOS",12000,2,2025,"Compra de mercancía febrero"),
    ("2025-02-05","Activos","ABONOS",12000,2,2025,"Pago mercancía en efectivo"),
    ("2025-02-07","Activos","CARGOS",11000,2,2025,"Cobro ventas semana 1 feb"),
    ("2025-02-07","Ingresos","ABONOS",11000,2,2025,"Ventas semana 1 febrero"),
    ("2025-02-14","Activos","CARGOS",13500,2,2025,"Cobro ventas San Valentín"),
    ("2025-02-14","Ingresos","ABONOS",13500,2,2025,"Ventas San Valentín"),
    ("2025-02-21","Activos","CARGOS",10200,2,2025,"Cobro ventas semana 3 feb"),
    ("2025-02-21","Ingresos","ABONOS",10200,2,2025,"Ventas semana 3 febrero"),
    ("2025-02-28","Activos","CARGOS", 9800,2,2025,"Cobro ventas semana 4 feb"),
    ("2025-02-28","Ingresos","ABONOS",9800,2,2025,"Ventas semana 4 febrero"),
    ("2025-02-28","Gastos", "CARGOS", 5000,2,2025,"Sueldo empleado febrero"),
    ("2025-02-28","Gastos", "CARGOS", 3500,2,2025,"Renta local febrero"),
    ("2025-02-28","Gastos", "CARGOS",  950,2,2025,"Servicios luz y agua febrero"),
    ("2025-02-28","Gastos", "CARGOS", 1200,2,2025,"Publicidad redes sociales"),
    ("2025-02-28","Activos","ABONOS",10650,2,2025,"Pago gastos operativos febrero"),
    ("2025-03-03","Gastos", "CARGOS",18000,3,2025,"Compra mercancía marzo a crédito"),
    ("2025-03-03","Pasivos","ABONOS",18000,3,2025,"Deuda por mercancía marzo"),
    ("2025-03-07","Activos","CARGOS", 9500,3,2025,"Cobro ventas semana 1 mar"),
    ("2025-03-07","Ingresos","ABONOS",9500,3,2025,"Ventas semana 1 marzo"),
    ("2025-03-14","Activos","CARGOS",10800,3,2025,"Cobro ventas semana 2 mar"),
    ("2025-03-14","Ingresos","ABONOS",10800,3,2025,"Ventas semana 2 marzo"),
    ("2025-03-15","Pasivos","CARGOS",10000,3,2025,"Abono a deuda de mercancía"),
    ("2025-03-15","Activos","ABONOS",10000,3,2025,"Pago deuda mercancía efectivo"),
    ("2025-03-21","Activos","CARGOS", 8900,3,2025,"Cobro ventas semana 3 mar"),
    ("2025-03-21","Ingresos","ABONOS",8900,3,2025,"Ventas semana 3 marzo"),
    ("2025-03-28","Activos","CARGOS", 7600,3,2025,"Cobro ventas semana 4 mar"),
    ("2025-03-28","Ingresos","ABONOS",7600,3,2025,"Ventas semana 4 marzo"),
    ("2025-03-31","Gastos", "CARGOS", 5500,3,2025,"Sueldo empleado marzo"),
    ("2025-03-31","Gastos", "CARGOS", 3500,3,2025,"Renta local marzo"),
    ("2025-03-31","Gastos", "CARGOS",  870,3,2025,"Servicios luz y agua marzo"),
    ("2025-03-31","Gastos", "CARGOS", 1500,3,2025,"Publicidad primavera"),
    ("2025-03-31","Gastos", "CARGOS",  600,3,2025,"Mantenimiento equipo"),
    ("2025-03-31","Activos","ABONOS",11970,3,2025,"Pago gastos operativos marzo"),
]


def _crear_workbook_vacio():
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"
    headers = ["Fecha","Cuenta","Tipo","Monto","Mes","Año","Descripcion"]
    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", start_color="1E3A5F")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    widths = [14, 14, 10, 12, 8, 8, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return wb


def init_excel():
    """Crea el archivo de trabajo; si no existe, lo inicializa con datos de prueba."""
    if not os.path.exists(WORK_FILE):
        wb = _crear_workbook_vacio()
        ws = wb.active
        thin = Border(bottom=Side(style="thin", color="2E3347"))
        for mov in DATOS_SEMILLA:
            fecha, cuenta, tipo, monto, mes, anio, desc = mov
            next_row = ws.max_row + 1
            row_data = [fecha, cuenta, tipo, monto, mes, anio, desc]
            cf = PatternFill("solid", start_color="3B1C1C")
            af = PatternFill("solid", start_color="1A3B2E")
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=next_row, column=col, value=val)
                cell.border = thin
                if col == 3:
                    cell.fill = cf if tipo == "CARGOS" else af
        wb.save(WORK_FILE)


def load_data() -> pd.DataFrame:
    init_excel()
    try:
        df = pd.read_excel(WORK_FILE, sheet_name="Movimientos")
        if df.empty or "Fecha" not in df.columns:
            return pd.DataFrame(columns=["Fecha","Cuenta","Tipo","Monto","Mes","Año","Descripcion"])
        df["Fecha"] = pd.to_datetime(df["Fecha"])
        df["Monto"] = pd.to_numeric(df["Monto"], errors="coerce").fillna(0)
        return df
    except Exception:
        return pd.DataFrame(columns=["Fecha","Cuenta","Tipo","Monto","Mes","Año","Descripcion"])


def save_to_excel(fecha, cuenta, tipo, monto, descripcion=""):
    init_excel()
    wb = load_workbook(WORK_FILE)
    ws = wb["Movimientos"]
    next_row = ws.max_row + 1
    dt = pd.to_datetime(fecha)
    row_data = [dt.strftime("%Y-%m-%d"), cuenta, tipo, monto, dt.month, dt.year, descripcion]
    cargo_fill = PatternFill("solid", start_color="3B1C1C")
    abono_fill = PatternFill("solid", start_color="1A3B2E")
    thin = Border(bottom=Side(style="thin", color="2E3347"))
    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=next_row, column=col, value=val)
        cell.border = thin
        if col == 3:
            cell.fill = cargo_fill if tipo == "CARGOS" else abono_fill
    wb.save(WORK_FILE)


# ─── LOGIC ────────────────────────────────────────────────────────────────────
def get_cuentas_list():
    if "cuentas_custom" not in st.session_state:
        st.session_state.cuentas_custom = []
    return CUENTAS_DEFAULT + st.session_state.cuentas_custom


def get_naturaleza(cuenta):
    return NATURALEZA.get(cuenta, "deudora")


def calcular_saldo(df: pd.DataFrame, cuenta: str) -> float:
    sub = df[df["Cuenta"] == cuenta]
    nat = get_naturaleza(cuenta)
    saldo = 0.0
    for _, row in sub.iterrows():
        if nat == "deudora":
            saldo += row["Monto"] if row["Tipo"] == "CARGOS" else -row["Monto"]
        else:
            saldo += row["Monto"] if row["Tipo"] == "ABONOS" else -row["Monto"]
    return saldo


def calcular_totales(df, cuenta):
    sub = df[df["Cuenta"] == cuenta]
    cargos = sub[sub["Tipo"] == "CARGOS"]["Monto"].sum()
    abonos = sub[sub["Tipo"] == "ABONOS"]["Monto"].sum()
    return cargos, abonos


def reset_todo():
    """Borra todos los movimientos, vuelve a cargar los datos de prueba."""
    if os.path.exists(WORK_FILE):
        os.remove(WORK_FILE)
    st.session_state.cuentas_custom = []
    for k in list(NATURALEZA.keys()):
        if k not in ["Activos", "Pasivos", "Capital", "Ingresos", "Gastos"]:
            del NATURALEZA[k]
    init_excel()  # recrea con datos semilla


def _reescribir_excel(df):
    """Escribe un DataFrame completo en WORK_FILE."""
    wb = _crear_workbook_vacio()
    ws = wb.active
    thin = Border(bottom=Side(style="thin", color="2E3347"))
    cf = PatternFill("solid", start_color="3B1C1C")
    af = PatternFill("solid", start_color="1A3B2E")
    for _, row in df.iterrows():
        dt = pd.to_datetime(row["Fecha"])
        row_data = [dt.strftime("%Y-%m-%d"), row["Cuenta"], row["Tipo"],
                    row["Monto"], row["Mes"], row["Año"], row.get("Descripcion","")]
        next_row = ws.max_row + 1
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=next_row, column=col, value=val)
            cell.border = thin
            if col == 3:
                cell.fill = cf if row["Tipo"] == "CARGOS" else af
    wb.save(WORK_FILE)


def eliminar_ultimo():
    df = load_data()
    if df.empty:
        return False
    _reescribir_excel(df.iloc[:-1])
    return True


def eliminar_por_indice(idx):
    df = load_data()
    if df.empty or idx >= len(df):
        return False
    _reescribir_excel(df.drop(df.index[idx]).reset_index(drop=True))
    return True


# ─── COMPONENTS ───────────────────────────────────────────────────────────────
def render_cuenta_t(cuenta, df):
    cargos_val, abonos_val = calcular_totales(df, cuenta)
    saldo = calcular_saldo(df, cuenta)
    nat = get_naturaleza(cuenta)
    color_saldo = "positive" if saldo >= 0 else "negative"

    sub = df[df["Cuenta"] == cuenta]
    cargos_items = sub[sub["Tipo"] == "CARGOS"]["Monto"].tolist()
    abonos_items = sub[sub["Tipo"] == "ABONOS"]["Monto"].tolist()

    cargos_html = "".join(f'<div class="t-amount">$ {v:,.2f}</div>' for v in cargos_items) or '<div style="color:#4b5563;font-size:0.8rem;">—</div>'
    abonos_html = "".join(f'<div class="t-amount">$ {v:,.2f}</div>' for v in abonos_items) or '<div style="color:#4b5563;font-size:0.8rem;">—</div>'

    naturaleza_label = "Nat. Deudora" if nat == "deudora" else "Nat. Acreedora"

    st.markdown(f"""
    <div class="cuenta-t">
      <div class="cuenta-t-header">
        <span>{cuenta}</span>
        <span style="font-size:0.7rem;color:#6b7280;">{naturaleza_label}</span>
      </div>
      <div class="cuenta-t-body">
        <div class="cuenta-t-col">
          <div class="cargo-label">Cargos</div>
          {cargos_html}
          <div style="margin-top:0.5rem;border-top:1px solid #2e3347;padding-top:0.4rem;font-size:0.8rem;color:#f87171;">
            Total: $ {cargos_val:,.2f}
          </div>
        </div>
        <div class="cuenta-t-divider"></div>
        <div class="cuenta-t-col">
          <div class="abono-label">Abonos</div>
          {abonos_html}
          <div style="margin-top:0.5rem;border-top:1px solid #2e3347;padding-top:0.4rem;font-size:0.8rem;color:#34d399;">
            Total: $ {abonos_val:,.2f}
          </div>
        </div>
      </div>
      <div style="padding:0.5rem 1rem;border-top:1px solid #2e3347;display:flex;justify-content:space-between;align-items:center;font-size:0.85rem;">
        <span style="color:#6b7280;">Saldo</span>
        <span class="{color_saldo}" style="font-weight:600;">$ {saldo:,.2f}</span>
      </div>
    </div>
    """, unsafe_allow_html=True)


# ─── SIDEBAR ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 📒 ContaT")
    st.markdown('<div style="font-size:0.75rem;color:#6b7280;margin-bottom:1.5rem;">Sistema Contable · Cuentas T</div>', unsafe_allow_html=True)

    st.markdown("### Nuevo Movimiento")
    with st.form("form_movimiento", clear_on_submit=True):
        cuentas_disponibles = get_cuentas_list()
        cuenta_sel = st.selectbox("Cuenta", cuentas_disponibles)
        tipo_sel = st.selectbox("Tipo de Movimiento", MOVIMIENTOS)
        monto_sel = st.text_input("Monto ($)", placeholder="Ej: 1500 o 1500.50")
        fecha_sel = st.date_input("Fecha", value=date.today())
        desc_sel = st.text_input("Descripción (opcional)")
        submitted = st.form_submit_button("➕ Registrar Movimiento", use_container_width=True)

    if submitted:
        try:
            monto_val = float(str(monto_sel).replace(",", ".").strip())
        except (ValueError, TypeError):
            monto_val = 0
        if monto_val > 0:
            save_to_excel(fecha_sel, cuenta_sel, tipo_sel, monto_val, desc_sel)
            st.success(f"✓ {tipo_sel} de ${monto_val:,.2f} en {cuenta_sel}")
            st.rerun()
        else:
            st.error("Ingresa un monto válido mayor a 0 (ej: 1500 o 1500.50)")

    st.divider()
    st.markdown("### Agregar Cuenta")

    # Catálogo de cuentas comunes con su naturaleza predefinida
    CATALOGO = {
        # ── ACTIVOS (deudora) ──
        "Caja": "deudora",
        "Bancos": "deudora",
        "Cuentas por cobrar": "deudora",
        "Inventario": "deudora",
        "Mercancías": "deudora",
        "Terrenos": "deudora",
        "Edificios": "deudora",
        "Equipo de cómputo": "deudora",
        "Equipo de transporte": "deudora",
        "Maquinaria": "deudora",
        "Mobiliario y equipo": "deudora",
        "Documentos por cobrar": "deudora",
        "IVA acreditable": "deudora",
        "Papelería y útiles": "deudora",
        "Seguros pagados por anticipado": "deudora",
        "Rentas pagadas por anticipado": "deudora",
        # ── PASIVOS (acreedora) ──
        "Cuentas por pagar": "acreedora",
        "Documentos por pagar": "acreedora",
        "Préstamos bancarios": "acreedora",
        "Acreedores diversos": "acreedora",
        "IVA por pagar": "acreedora",
        "ISR por pagar": "acreedora",
        "Sueldos por pagar": "acreedora",
        "Hipotecas por pagar": "acreedora",
        "Intereses por pagar": "acreedora",
        # ── CAPITAL (acreedora) ──
        "Capital social": "acreedora",
        "Utilidad del ejercicio": "acreedora",
        "Pérdida del ejercicio": "deudora",
        "Reserva legal": "acreedora",
        "Dividendos": "deudora",
        # ── INGRESOS (acreedora) ──
        "Ventas": "acreedora",
        "Servicios prestados": "acreedora",
        "Intereses ganados": "acreedora",
        "Comisiones ganadas": "acreedora",
        "Descuentos sobre compras": "acreedora",
        "Devoluciones sobre ventas": "deudora",
        # ── GASTOS (deudora) ──
        "Sueldos y salarios": "deudora",
        "Renta": "deudora",
        "Luz y agua": "deudora",
        "Teléfono e internet": "deudora",
        "Publicidad": "deudora",
        "Compras": "deudora",
        "Fletes sobre compras": "deudora",
        "Descuentos sobre ventas": "deudora",
        "Devoluciones sobre compras": "acreedora",
        "Depreciación": "deudora",
        "Intereses pagados": "deudora",
        "Mantenimiento": "deudora",
        "Seguros": "deudora",
        "Papelería": "deudora",
        "Combustibles": "deudora",
    }

    # Filtrar las que ya están agregadas
    ya_agregadas = set(get_cuentas_list())
    catalogo_disponible = {k: v for k, v in CATALOGO.items() if k not in ya_agregadas}

    # Selectbox del catálogo + opción personalizada
    opciones_catalogo = ["— Selecciona una cuenta —"] + sorted(catalogo_disponible.keys()) + ["✏️ Escribir nombre personalizado"]
    cuenta_elegida = st.selectbox("Cuenta a agregar", opciones_catalogo, key="sel_nueva_cuenta")

    if cuenta_elegida == "✏️ Escribir nombre personalizado":
        nombre_custom = st.text_input("Nombre personalizado", placeholder="Ej: Inversiones temporales")
        nat_custom = st.selectbox("Naturaleza", ["deudora", "acreedora"], key="nat_custom")
        if st.button("➕ Agregar Cuenta", use_container_width=True, key="btn_add_custom"):
            nombre = nombre_custom.strip().capitalize()
            if nombre and nombre not in get_cuentas_list():
                st.session_state.cuentas_custom.append(nombre)
                NATURALEZA[nombre] = nat_custom
                st.success(f"✓ '{nombre}' añadida como cuenta {nat_custom}")
                st.rerun()
            elif not nombre:
                st.error("Escribe un nombre válido")
            else:
                st.warning("Esa cuenta ya existe")

    elif cuenta_elegida != "— Selecciona una cuenta —":
        nat_det = CATALOGO[cuenta_elegida]
        nat_color = "#93c5fd" if nat_det == "deudora" else "#34d399"
        nat_icon  = "📘" if nat_det == "deudora" else "📗"
        st.markdown(f"""
        <div style="background:#1f2438;border:1px solid {nat_color}44;border-left:3px solid {nat_color};
            border-radius:8px;padding:0.7rem 1rem;margin:0.4rem 0 0.6rem;">
            <div style="font-size:0.7rem;color:#6b7280;letter-spacing:0.1em;margin-bottom:0.2rem;">NATURALEZA DETECTADA</div>
            <div style="color:{nat_color};font-weight:700;font-size:0.9rem;">{nat_icon} {nat_det.upper()}</div>
            <div style="font-size:0.72rem;color:#9ca3af;margin-top:0.2rem;">
                {"Aumenta con Cargos · Disminuye con Abonos" if nat_det == "deudora" else "Aumenta con Abonos · Disminuye con Cargos"}
            </div>
        </div>
        """, unsafe_allow_html=True)
        if st.button(f"➕ Agregar '{cuenta_elegida}'", use_container_width=True, key="btn_add_catalogo"):
            st.session_state.cuentas_custom.append(cuenta_elegida)
            NATURALEZA[cuenta_elegida] = nat_det
            st.success(f"✓ '{cuenta_elegida}' añadida ({nat_det})")
            st.rerun()

    if st.session_state.get("cuentas_custom"):
        st.markdown("**Cuentas personalizadas:**")
        for i, c in enumerate(st.session_state.cuentas_custom):
            col_c, col_x = st.columns([3, 1])
            with col_c:
                st.markdown(f"· {c} ({NATURALEZA.get(c,'—')})")
            with col_x:
                if st.button("✕", key=f"del_cuenta_{i}", help=f"Eliminar {c}"):
                    st.session_state.cuentas_custom.pop(i)
                    if c in NATURALEZA:
                        del NATURALEZA[c]
                    st.rerun()

    st.divider()
    st.markdown("### ⚠️ Zona de Reset")

    if st.button("↩ Deshacer último movimiento", use_container_width=True):
        st.session_state["confirm_undo"] = True

    if st.session_state.get("confirm_undo"):
        st.warning("¿Eliminar el último movimiento registrado?")
        col_y, col_n = st.columns(2)
        with col_y:
            if st.button("✓ Sí", key="undo_yes", use_container_width=True):
                ok = eliminar_ultimo()
                st.session_state["confirm_undo"] = False
                if ok:
                    st.success("Último movimiento eliminado")
                else:
                    st.info("No hay movimientos para eliminar")
                st.rerun()
        with col_n:
            if st.button("✗ No", key="undo_no", use_container_width=True):
                st.session_state["confirm_undo"] = False
                st.rerun()

    st.markdown("")

    if st.button("🗑 Resetear TODO a cero", use_container_width=True):
        st.session_state["confirm_reset"] = True

    if st.session_state.get("confirm_reset"):
        st.error("¿Borrar TODOS los movimientos y cuentas personalizadas?")
        col_y2, col_n2 = st.columns(2)
        with col_y2:
            if st.button("✓ Sí, resetear", key="reset_yes", use_container_width=True):
                reset_todo()
                st.session_state["confirm_reset"] = False
                st.success("✓ Todo reseteado a cero")
                st.rerun()
        with col_n2:
            if st.button("✗ Cancelar", key="reset_no", use_container_width=True):
                st.session_state["confirm_reset"] = False
                st.rerun()


# ─── MAIN ─────────────────────────────────────────────────────────────────────
df_all = load_data()

st.markdown("# ContaT")
st.markdown('<div class="tab-header">Sistema de Cuentas T · Contabilidad Financiera</div>', unsafe_allow_html=True)

tab1, tab2, tab3, tab4 = st.tabs(["📊 Dashboard", "🧾 Cuentas T", "📈 Análisis", "📋 Historial"])

# ══════════════════════════════════════════════════════════════════════
# TAB 1 · DASHBOARD
# ══════════════════════════════════════════════════════════════════════
with tab1:
    cuentas_list = get_cuentas_list()

    # Métricas principales
    cols = st.columns(5)
    labels = ["Activos", "Pasivos", "Capital", "Ingresos", "Gastos"]
    colors = ["neutral", "negative", "positive", "positive", "negative"]
    for i, (col, lbl, clr) in enumerate(zip(cols, labels, colors)):
        saldo = calcular_saldo(df_all, lbl) if not df_all.empty else 0
        with col:
            st.markdown(f"""
            <div class="metric-card">
              <div class="label">{lbl}</div>
              <div class="value {clr}">$ {saldo:,.0f}</div>
            </div>
            """, unsafe_allow_html=True)

    st.markdown("---")

    # ESF
    activos = calcular_saldo(df_all, "Activos")
    pasivos = calcular_saldo(df_all, "Pasivos")
    capital = calcular_saldo(df_all, "Capital")
    cuadrado = abs(activos - (pasivos + capital)) < 0.01

    col_esf, col_res = st.columns([1, 1])
    with col_esf:
        if cuadrado:
            st.markdown(f"""
            <div class="esf-ok">
              ✓ ESF Cuadrado · Activos = Pasivos + Capital<br>
              <span style="font-size:0.85rem;opacity:0.8;">
                $ {activos:,.2f} = $ {pasivos:,.2f} + $ {capital:,.2f}
              </span>
            </div>""", unsafe_allow_html=True)
        else:
            diff = activos - (pasivos + capital)
            st.markdown(f"""
            <div class="esf-fail">
              ✗ ESF No cuadrado · Diferencia: $ {diff:,.2f}<br>
              <span style="font-size:0.85rem;opacity:0.8;">
                Activos: $ {activos:,.2f} ≠ Pasivos + Capital: $ {pasivos + capital:,.2f}
              </span>
            </div>""", unsafe_allow_html=True)

    with col_res:
        ingresos = calcular_saldo(df_all, "Ingresos")
        gastos = calcular_saldo(df_all, "Gastos")
        utilidad = ingresos - gastos
        color_u = "positive" if utilidad >= 0 else "negative"
        tipo_u = "Utilidad Neta" if utilidad >= 0 else "Pérdida Neta"
        st.markdown(f"""
        <div class="metric-card">
          <div class="label">{tipo_u}</div>
          <div class="value {color_u}">$ {utilidad:,.2f}</div>
          <div style="font-size:0.75rem;color:#6b7280;margin-top:0.4rem;">
            Ingresos $ {ingresos:,.2f} − Gastos $ {gastos:,.2f}
          </div>
        </div>""", unsafe_allow_html=True)

    # Mini gráfica de saldos
    if not df_all.empty:
        saldos_data = {c: calcular_saldo(df_all, c) for c in cuentas_list}
        fig_bar = px.bar(
            x=list(saldos_data.keys()),
            y=list(saldos_data.values()),
            labels={"x": "Cuenta", "y": "Saldo ($)"},
            color=list(saldos_data.values()),
            color_continuous_scale=["#f87171", "#fbbf24", "#34d399"],
        )
        fig_bar.update_layout(
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
            font_color="#e8e3d5",
            showlegend=False,
            coloraxis_showscale=False,
            margin=dict(l=10, r=10, t=10, b=10),
            height=280,
            xaxis=dict(gridcolor="#2e3347"),
            yaxis=dict(gridcolor="#2e3347"),
        )
        st.plotly_chart(fig_bar, use_container_width=True)

    # ── SECCIÓN EDUCATIVA ──────────────────────────────────────────────
    st.markdown("---")
    st.markdown("## 📚 Guía de Conceptos Contables")
    st.markdown('<div style="font-size:0.8rem;color:#6b7280;margin-bottom:1.5rem;">Referencia rápida para entender los elementos del sistema contable</div>', unsafe_allow_html=True)

    # Tarjetas de los 5 tipos de cuenta
    st.markdown("### 🗂 Tipos de Cuenta")
    info_cuentas = [
        {
            "nombre": "ACTIVOS",
            "icono": "🏦",
            "color_borde": "#93c5fd",
            "color_icon_bg": "#1e3a5f",
            "definicion": "Todo lo que la empresa <b>posee o tiene derecho a recibir</b>. Son los recursos económicos con los que cuenta el negocio.",
            "ejemplos": "Efectivo · Cuentas por cobrar · Inventario · Equipo · Edificios · Vehículos",
            "naturaleza": "deudora",
            "regla": "Aumentan con CARGOS · Disminuyen con ABONOS",
        },
        {
            "nombre": "PASIVOS",
            "icono": "📋",
            "color_borde": "#f87171",
            "color_icon_bg": "#3b1c1c",
            "definicion": "Todo lo que la empresa <b>debe a terceros</b>. Son las obligaciones o deudas con acreedores externos.",
            "ejemplos": "Préstamos bancarios · Cuentas por pagar · Impuestos por pagar · Hipotecas",
            "naturaleza": "acreedora",
            "regla": "Aumentan con ABONOS · Disminuyen con CARGOS",
        },
        {
            "nombre": "CAPITAL",
            "icono": "💼",
            "color_borde": "#34d399",
            "color_icon_bg": "#1a3b2e",
            "definicion": "La <b>inversión de los dueños</b> en la empresa. Es la diferencia entre Activos y Pasivos (patrimonio neto).",
            "ejemplos": "Aportaciones del dueño · Utilidades retenidas · Reservas · Capital social",
            "naturaleza": "acreedora",
            "regla": "Aumentan con ABONOS · Disminuyen con CARGOS",
        },
        {
            "nombre": "INGRESOS",
            "icono": "📈",
            "color_borde": "#fbbf24",
            "color_icon_bg": "#3b2e0a",
            "definicion": "Dinero que la empresa <b>gana por sus operaciones</b>. Incrementan el capital del negocio.",
            "ejemplos": "Ventas · Servicios prestados · Intereses ganados · Comisiones · Rentas cobradas",
            "naturaleza": "acreedora",
            "regla": "Aumentan con ABONOS · Disminuyen con CARGOS",
        },
        {
            "nombre": "GASTOS",
            "icono": "📉",
            "color_borde": "#c084fc",
            "color_icon_bg": "#2e1a47",
            "definicion": "Dinero que la empresa <b>gasta para operar</b>. Reducen el capital al consumir recursos.",
            "ejemplos": "Sueldos · Renta · Servicios · Publicidad · Papelería · Depreciación",
            "naturaleza": "deudora",
            "regla": "Aumentan con CARGOS · Disminuyen con ABONOS",
        },
    ]

    col_a, col_b, col_c = st.columns(3)
    col_d, col_e, _ = st.columns(3)
    columnas_edu = [col_a, col_b, col_c, col_d, col_e]

    for col, info in zip(columnas_edu, info_cuentas):
        nat_color = "#93c5fd" if info["naturaleza"] == "deudora" else "#34d399"
        nat_bg    = "#1e3a5f" if info["naturaleza"] == "deudora" else "#1a3b2e"
        nat_label = "Nat. Deudora" if info["naturaleza"] == "deudora" else "Nat. Acreedora"
        with col:
            st.markdown(f"""
            <div style="
                background: linear-gradient(160deg, #1a1e2e, #1f2438);
                border: 1px solid {info['color_borde']}44;
                border-top: 3px solid {info['color_borde']};
                border-radius: 12px;
                padding: 1.2rem;
                margin-bottom: 1rem;
                height: 100%;
            ">
                <div style="display:flex;align-items:center;gap:0.6rem;margin-bottom:0.8rem;">
                    <div style="background:{info['color_icon_bg']};border-radius:8px;padding:0.4rem 0.6rem;font-size:1.3rem;">
                        {info['icono']}
                    </div>
                    <div>
                        <div style="font-size:0.65rem;letter-spacing:0.12em;color:#6b7280;">CUENTA</div>
                        <div style="font-family:'DM Serif Display',serif;font-size:1rem;color:{info['color_borde']};">{info['nombre']}</div>
                    </div>
                </div>
                <p style="font-size:0.82rem;color:#d1d5db;line-height:1.5;margin-bottom:0.8rem;">
                    {info['definicion']}
                </p>
                <div style="background:#0f1117;border-radius:6px;padding:0.5rem 0.7rem;margin-bottom:0.7rem;">
                    <div style="font-size:0.65rem;color:#6b7280;letter-spacing:0.1em;margin-bottom:0.3rem;">EJEMPLOS</div>
                    <div style="font-size:0.75rem;color:#9ca3af;">{info['ejemplos']}</div>
                </div>
                <div style="display:flex;justify-content:space-between;align-items:center;">
                    <span style="background:{nat_bg};color:{nat_color};font-size:0.65rem;font-weight:700;
                        letter-spacing:0.08em;padding:2px 8px;border-radius:20px;border:1px solid {nat_color}44;">
                        {nat_label}
                    </span>
                    <span style="font-size:0.68rem;color:#6b7280;">{info['regla']}</span>
                </div>
            </div>
            """, unsafe_allow_html=True)

    # Naturalezas: Deudora vs Acreedora
    st.markdown("### ⚖️ Naturaleza de las Cuentas")
    col_deu, col_acre = st.columns(2)

    with col_deu:
        st.markdown("""
        <div style="
            background: linear-gradient(135deg, #1e3a5f22, #1f2438);
            border: 1px solid #93c5fd44;
            border-left: 4px solid #93c5fd;
            border-radius: 10px;
            padding: 1.2rem 1.4rem;
        ">
            <div style="font-family:'DM Serif Display',serif;font-size:1.1rem;color:#93c5fd;margin-bottom:0.6rem;">
                📘 Naturaleza Deudora
            </div>
            <p style="font-size:0.83rem;color:#d1d5db;line-height:1.6;margin-bottom:0.8rem;">
                Las cuentas de naturaleza deudora <b>aumentan su saldo cuando se cargan (lado izquierdo de la T)</b>
                y disminuyen cuando se abonan.
            </p>
            <div style="background:#0f1117;border-radius:8px;padding:0.8rem;margin-bottom:0.6rem;">
                <div style="font-size:0.7rem;color:#6b7280;margin-bottom:0.4rem;letter-spacing:0.1em;">CUENTAS CON ESTA NATURALEZA</div>
                <div style="display:flex;gap:0.5rem;flex-wrap:wrap;">
                    <span style="background:#1e3a5f;color:#93c5fd;padding:3px 10px;border-radius:20px;font-size:0.75rem;">🏦 Activos</span>
                    <span style="background:#2e1a47;color:#c084fc;padding:3px 10px;border-radius:20px;font-size:0.75rem;">📉 Gastos</span>
                </div>
            </div>
            <div style="font-size:0.78rem;color:#9ca3af;line-height:1.5;">
                💡 <i>Su saldo normal es <b style="color:#93c5fd;">positivo</b> en el lado del Cargo (izquierdo).</i>
            </div>
        </div>
        """, unsafe_allow_html=True)

    with col_acre:
        st.markdown("""
        <div style="
            background: linear-gradient(135deg, #1a3b2e22, #1f2438);
            border: 1px solid #34d39944;
            border-left: 4px solid #34d399;
            border-radius: 10px;
            padding: 1.2rem 1.4rem;
        ">
            <div style="font-family:'DM Serif Display',serif;font-size:1.1rem;color:#34d399;margin-bottom:0.6rem;">
                📗 Naturaleza Acreedora
            </div>
            <p style="font-size:0.83rem;color:#d1d5db;line-height:1.6;margin-bottom:0.8rem;">
                Las cuentas de naturaleza acreedora <b>aumentan su saldo cuando se abonan (lado derecho de la T)</b>
                y disminuyen cuando se cargan.
            </p>
            <div style="background:#0f1117;border-radius:8px;padding:0.8rem;margin-bottom:0.6rem;">
                <div style="font-size:0.7rem;color:#6b7280;margin-bottom:0.4rem;letter-spacing:0.1em;">CUENTAS CON ESTA NATURALEZA</div>
                <div style="display:flex;gap:0.5rem;flex-wrap:wrap;">
                    <span style="background:#3b1c1c;color:#f87171;padding:3px 10px;border-radius:20px;font-size:0.75rem;">📋 Pasivos</span>
                    <span style="background:#1a3b2e;color:#34d399;padding:3px 10px;border-radius:20px;font-size:0.75rem;">💼 Capital</span>
                    <span style="background:#3b2e0a;color:#fbbf24;padding:3px 10px;border-radius:20px;font-size:0.75rem;">📈 Ingresos</span>
                </div>
            </div>
            <div style="font-size:0.78rem;color:#9ca3af;line-height:1.5;">
                💡 <i>Su saldo normal es <b style="color:#34d399;">positivo</b> en el lado del Abono (derecho).</i>
            </div>
        </div>
        """, unsafe_allow_html=True)

    # Diagrama visual de la Cuenta T
    st.markdown("### 🔠 ¿Cómo funciona la Cuenta T?")
    st.markdown("""
    <div style="
        background: #161922;
        border: 1px solid #2e3347;
        border-radius: 12px;
        padding: 1.4rem;
    ">
        <p style="font-size:0.85rem;color:#9ca3af;margin-bottom:1rem;">
            La <b style="color:#e8e3d5;">Cuenta T</b> es la representación gráfica de una cuenta contable.
            Su nombre viene de su forma: una línea vertical divide el <b style="color:#f87171;">Cargo (izquierda)</b>
            del <b style="color:#34d399;">Abono (derecha)</b>.
        </p>
        <div style="display:grid;grid-template-columns:1fr 1fr;max-width:480px;margin:0 auto;
            border:1px solid #2e3347;border-radius:8px;overflow:hidden;">
            <div style="background:#1e3a5f;padding:0.5rem;text-align:center;
                font-family:'DM Serif Display',serif;font-size:1rem;color:#93c5fd;
                border-bottom:2px solid #93c5fd;grid-column:1/-1;">
                Nombre de la Cuenta
            </div>
            <div style="padding:0.8rem;border-right:2px solid #2e3347;">
                <div style="color:#f87171;font-size:0.7rem;letter-spacing:0.1em;font-weight:700;margin-bottom:0.5rem;">
                    DEBE / CARGO
                </div>
                <div style="color:#9ca3af;font-size:0.78rem;line-height:1.6;">
                    · Aumenta Activos<br>
                    · Aumenta Gastos<br>
                    · Disminuye Pasivos<br>
                    · Disminuye Capital<br>
                    · Disminuye Ingresos
                </div>
            </div>
            <div style="padding:0.8rem;">
                <div style="color:#34d399;font-size:0.7rem;letter-spacing:0.1em;font-weight:700;margin-bottom:0.5rem;">
                    HABER / ABONO
                </div>
                <div style="color:#9ca3af;font-size:0.78rem;line-height:1.6;">
                    · Disminuye Activos<br>
                    · Disminuye Gastos<br>
                    · Aumenta Pasivos<br>
                    · Aumenta Capital<br>
                    · Aumenta Ingresos
                </div>
            </div>
        </div>
        <div style="margin-top:1rem;background:#0f1117;border-radius:8px;padding:0.8rem 1rem;
            border-left:3px solid #fbbf24;">
            <span style="color:#fbbf24;font-weight:700;font-size:0.8rem;">📌 Regla de oro: </span>
            <span style="color:#9ca3af;font-size:0.82rem;">
                En toda transacción el total de Cargos siempre debe ser igual al total de Abonos.
                Esto garantiza que la ecuación contable <b style="color:#e8e3d5;">Activo = Pasivo + Capital</b> siempre esté en equilibrio.
            </span>
        </div>
    </div>
    """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════
# TAB 2 · CUENTAS T
# ══════════════════════════════════════════════════════════════════════
with tab2:
    cuentas_list = get_cuentas_list()
    cols_per_row = 2
    rows = [cuentas_list[i:i+cols_per_row] for i in range(0, len(cuentas_list), cols_per_row)]
    for row in rows:
        cols = st.columns(cols_per_row)
        for col, cuenta in zip(cols, row):
            with col:
                render_cuenta_t(cuenta, df_all)


# ══════════════════════════════════════════════════════════════════════
# TAB 3 · ANÁLISIS
# ══════════════════════════════════════════════════════════════════════
with tab3:
    if df_all.empty:
        st.info("No hay datos registrados aún. Agrega movimientos desde el panel lateral.")
    else:
        col_ctrl1, col_ctrl2 = st.columns([1, 2])
        with col_ctrl1:
            periodo = st.selectbox("Ver por", ["Mes", "Año"])
        with col_ctrl2:
            cuentas_filtro = st.multiselect(
                "Filtrar cuentas",
                get_cuentas_list(),
                default=["Ingresos", "Gastos", "Activos"],
            )

        df_analysis = df_all.copy()

        MESES = {1:"Ene",2:"Feb",3:"Mar",4:"Abr",5:"May",6:"Jun",
                 7:"Jul",8:"Ago",9:"Sep",10:"Oct",11:"Nov",12:"Dic"}

        if periodo == "Mes":
            df_analysis["Periodo"] = df_analysis["Mes"].map(MESES) + " " + df_analysis["Año"].astype(str)
            df_analysis["sort_key"] = df_analysis["Año"] * 100 + df_analysis["Mes"]
        else:
            df_analysis["Periodo"] = df_analysis["Año"].astype(str)
            df_analysis["sort_key"] = df_analysis["Año"]

        if cuentas_filtro:
            df_filtered = df_analysis[df_analysis["Cuenta"].isin(cuentas_filtro)]
        else:
            df_filtered = df_analysis.copy()

        # Calcular saldo neto por período y cuenta según naturaleza
        def saldo_neto(row):
            nat = get_naturaleza(row["Cuenta"])
            if nat == "deudora":
                return row["Monto"] if row["Tipo"] == "CARGOS" else -row["Monto"]
            else:
                return row["Monto"] if row["Tipo"] == "ABONOS" else -row["Monto"]

        df_filtered = df_filtered.copy()
        df_filtered["Saldo"] = df_filtered.apply(saldo_neto, axis=1)

        pivot = df_filtered.groupby(["sort_key", "Periodo", "Cuenta"])["Saldo"].sum().reset_index()
        pivot = pivot.sort_values("sort_key")

        # Gráfico de líneas por cuenta
        fig_line = px.line(
            pivot,
            x="Periodo",
            y="Saldo",
            color="Cuenta",
            markers=True,
            title=f"Evolución de Saldos por {periodo}",
            color_discrete_sequence=["#93c5fd","#34d399","#f87171","#fbbf24","#c084fc","#fb923c"],
        )
        fig_line.update_layout(
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
            font_color="#e8e3d5",
            title_font_family="DM Serif Display",
            legend=dict(bgcolor="rgba(0,0,0,0)"),
            xaxis=dict(gridcolor="#2e3347"),
            yaxis=dict(gridcolor="#2e3347"),
            height=350,
        )
        st.plotly_chart(fig_line, use_container_width=True)

        col_a, col_b = st.columns(2)

        with col_a:
            # Barras apiladas Cargos vs Abonos
            df_tipo = df_filtered.groupby(["Periodo", "sort_key", "Tipo"])["Monto"].sum().reset_index().sort_values("sort_key")
            fig_stack = px.bar(
                df_tipo,
                x="Periodo",
                y="Monto",
                color="Tipo",
                barmode="group",
                title="Cargos vs Abonos",
                color_discrete_map={"CARGOS": "#f87171", "ABONOS": "#34d399"},
            )
            fig_stack.update_layout(
                paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor="rgba(0,0,0,0)",
                font_color="#e8e3d5",
                title_font_family="DM Serif Display",
                legend=dict(bgcolor="rgba(0,0,0,0)"),
                xaxis=dict(gridcolor="#2e3347"),
                yaxis=dict(gridcolor="#2e3347"),
                height=300,
            )
            st.plotly_chart(fig_stack, use_container_width=True)

        with col_b:
            # Pie distribución por cuenta
            pie_data = df_filtered.groupby("Cuenta")["Monto"].sum().reset_index()
            fig_pie = px.pie(
                pie_data,
                values="Monto",
                names="Cuenta",
                title="Distribución por Cuenta",
                color_discrete_sequence=["#93c5fd","#34d399","#f87171","#fbbf24","#c084fc","#fb923c"],
                hole=0.4,
            )
            fig_pie.update_layout(
                paper_bgcolor="rgba(0,0,0,0)",
                font_color="#e8e3d5",
                title_font_family="DM Serif Display",
                legend=dict(bgcolor="rgba(0,0,0,0)"),
                height=300,
            )
            st.plotly_chart(fig_pie, use_container_width=True)

        # Tabla resumen
        st.markdown("#### Resumen por Período")
        tabla_resumen = pivot.pivot_table(
            index="Periodo", columns="Cuenta", values="Saldo", aggfunc="sum", fill_value=0
        )
        tabla_resumen.index.name = "Período"
        st.dataframe(
            tabla_resumen.style.format("${:,.2f}"),
            use_container_width=True,
        )

        # Insights
        st.markdown("#### 🔍 Insights")
        if "Ingresos" in df_all["Cuenta"].values and "Gastos" in df_all["Cuenta"].values:
            df_ig = df_analysis[df_analysis["Cuenta"].isin(["Ingresos","Gastos"])].copy()
            df_ig["Saldo"] = df_ig.apply(saldo_neto, axis=1)
            if periodo == "Mes":
                top = df_ig[df_ig["Cuenta"]=="Ingresos"].groupby("Periodo")["Saldo"].sum()
                if not top.empty:
                    mes_max = top.idxmax()
                    st.success(f"📈 El período con más ingresos fue **{mes_max}** con **${top.max():,.2f}**")
                bottom = df_ig[df_ig["Cuenta"]=="Gastos"].groupby("Periodo")["Saldo"].sum()
                if not bottom.empty:
                    mes_gasto = bottom.idxmax()
                    st.warning(f"📉 El período con más gastos fue **{mes_gasto}** con **${bottom.max():,.2f}**")


# ══════════════════════════════════════════════════════════════════════
# TAB 4 · HISTORIAL
# ══════════════════════════════════════════════════════════════════════
with tab4:
    if df_all.empty:
        st.info("Sin movimientos registrados.")
    else:
        col_f1, col_f2, col_f3 = st.columns(3)
        with col_f1:
            cuentas_h = ["Todas"] + get_cuentas_list()
            filtro_cuenta = st.selectbox("Filtrar por cuenta", cuentas_h)
        with col_f2:
            filtro_tipo = st.selectbox("Filtrar por tipo", ["Todos", "CARGOS", "ABONOS"])
        with col_f3:
            años = sorted(df_all["Año"].dropna().unique().astype(int).tolist(), reverse=True)
            filtro_año = st.selectbox("Año", ["Todos"] + [str(a) for a in años])

        df_hist = df_all.copy()
        if filtro_cuenta != "Todas":
            df_hist = df_hist[df_hist["Cuenta"] == filtro_cuenta]
        if filtro_tipo != "Todos":
            df_hist = df_hist[df_hist["Tipo"] == filtro_tipo]
        if filtro_año != "Todos":
            df_hist = df_hist[df_hist["Año"] == int(filtro_año)]

        df_hist = df_hist.sort_values("Fecha", ascending=False)

        def color_tipo(val):
            if val == "CARGOS":
                return "color: #f87171"
            elif val == "ABONOS":
                return "color: #34d399"
            return ""

        # Tabla con botón de eliminar por fila
        st.markdown("##### Movimientos registrados")
        col_h = st.columns([2, 2, 2, 2, 3, 1])
        for h, label in zip(col_h, ["Fecha", "Cuenta", "Tipo", "Monto", "Descripción", "Eliminar"]):
            h.markdown(f"**{label}**")

        df_hist_indexed = df_hist.reset_index()  # guarda índice original en columna 'index'
        for _, row in df_hist_indexed.iterrows():
            c1, c2, c3, c4, c5, c6 = st.columns([2, 2, 2, 2, 3, 1])
            c1.write(str(row["Fecha"])[:10])
            c2.write(row["Cuenta"])
            color = "#f87171" if row["Tipo"] == "CARGOS" else "#34d399"
            c3.markdown(f'<span style="color:{color};font-weight:600">{row["Tipo"]}</span>', unsafe_allow_html=True)
            c4.write(f"${row['Monto']:,.2f}")
            c5.write(row.get("Descripcion", "") or "—")
            if c6.button("✕", key=f"del_row_{row['index']}"):
                st.session_state["confirm_del_idx"] = int(row["index"])

        if "confirm_del_idx" in st.session_state and st.session_state["confirm_del_idx"] is not None:
            idx_to_del = st.session_state["confirm_del_idx"]
            st.warning(f"¿Eliminar el movimiento seleccionado (fila #{idx_to_del + 1})?")
            col_cd1, col_cd2, _ = st.columns([1, 1, 4])
            with col_cd1:
                if st.button("✓ Confirmar", key="confirm_del_btn"):
                    eliminar_por_indice(idx_to_del)
                    st.session_state["confirm_del_idx"] = None
                    st.rerun()
            with col_cd2:
                if st.button("✗ Cancelar", key="cancel_del_btn"):
                    st.session_state["confirm_del_idx"] = None
                    st.rerun()

        st.markdown(f"**{len(df_hist)} movimientos** · Total cargos: **${df_hist[df_hist['Tipo']=='CARGOS']['Monto'].sum():,.2f}** · Total abonos: **${df_hist[df_hist['Tipo']=='ABONOS']['Monto'].sum():,.2f}**")

        col_dl1, col_dl2 = st.columns([1, 4])
        with col_dl1:
            csv_data = df_hist.to_csv(index=False).encode("utf-8")
            st.download_button(
                "⬇ Descargar CSV",
                data=csv_data,
                file_name="historial_contable.csv",
                mime="text/csv",
                use_container_width=True,
            )
